from openpyxl import Workbook, styles
from openpyxl.chart import LineChart,Reference
from openpyxl.chart.layout import Layout, ManualLayout

from openpyxl.chart.axis import DateAxis
import requests
import os
import datetime

###------- CREATING SHEET AND STYLES ---------###

wb = Workbook()
ws = wb.active
ws.title = "Apple"
ws['A1'] = 'Apple Data'
title_style = styles.Font(bold=True, sz=30, underline='single', italic=True)
ws['A1'].font = title_style
tesla_tab = wb.create_sheet("Tesla")
tesla_tab['A1'] = "Tesla Data"
tesla_tab['A1'].font = title_style

###------- API VARIABLES ---------###

STOCK = "AAPL"

stock_api_key = 'P80S76BFMQ83FJBH'
STOCK_ENDPOINT = "https://www.alphavantage.co/query"

stock_param = {
    "function": "TIME_SERIES_DAILY_ADJUSTED",
    "symbol": STOCK,
    "apikey": stock_api_key,
}

#------- STOCK API CALLS ---------#

response = requests.get(STOCK_ENDPOINT, params=stock_param)
response.raise_for_status()
stock_data = response.json()["Time Series (Daily)"]

#------- RUN STOCK REQUESTS ---------#

apple_daily_prices = []

def value_check():
    for i in stock_data.keys():
        date = datetime.datetime.strptime(i,'%Y-%m-%d').date() # converts into date time object
        high_price = float((stock_data[i]['2. high']))
        low_price = float((stock_data[i]['3. low']))
        apple_daily_prices.append([date, high_price, low_price])

value_check()

#------ HANDLE 4:1 STOCK SPLIT -----#

for i in apple_daily_prices:
    if i[0] <= datetime.date(2020, 8, 28): #this is accessing the date time object from the new daily price list
        i[1] /= 4  # this is dividing low and high prices by 4 for the old un-adjusted prices
        i[2] /= 4

#------- IMPORTING PRICES INTO EXCEL ---------#

rows = [
    ['Date', 'Low', 'High']]

for row in rows:
    ws.append(row)
    for i in range(len(apple_daily_prices)):
        ws.append(apple_daily_prices[i])

#------- CREATING CHART ---------#

c1 = LineChart()
c1.title = "Apple Highs vs Lows"
c1.style = 13
c1.y_axis.scaling.min = 100
c1.y_axis.scaling.max = 140
c1.y_axis.title = 'Price'
c1.x_axis.title = 'Date'
c1.height = 15
c1.width = 30

data = Reference(ws, min_col=2, min_row=2, max_col=4, max_row=105)
c1.add_data(data, titles_from_data=True)

# Style the lines
s1 = c1.series[1]
s1.marker.symbol = "triangle"
s1.marker.graphicalProperties.solidFill = "FF0000" # Marker filling
s1.marker.graphicalProperties.line.solidFill = "FF0000" # Marker outline

s1.graphicalProperties.line.noFill = True

s2 = c1.series[2]
s2.graphicalProperties.line.solidFill = "00AAAA"
s2.graphicalProperties.line.dashStyle = "sysDot"
s2.graphicalProperties.line.width = 100050 # width in EMUs

s2 = c1.series[2]
s2.smooth = True # Make the line smooth


ws.add_chart(c1, "E05")

wb.save('stock_tracker.xlsx')

os.system('open stock_tracker.xlsx') # opens excel, can disable
print("done")

#TODO: Convert into OOP framework to allow multiple stocks and sheets
#TODO: Add date axis to the chart
#TODO: Create a function out of split functionality
#TODO: Create try/except module for API failure scenario
#TODO: Save to memory or access old data from spreadsheet if API fails
